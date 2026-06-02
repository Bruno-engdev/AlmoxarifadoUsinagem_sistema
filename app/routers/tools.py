"""
Tools router – CRUD, stock add/remove, search, tool registration, XLSX export.
"""

from io import BytesIO
from datetime import datetime
from fastapi import APIRouter, Request, Depends, Form, Query, HTTPException
from fastapi.responses import RedirectResponse
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import case, and_

from app.database import get_db
from app.models import Tool, ToolType, ToolParameter, Employee, Machine
from app.auth import require_login
from app.pagination import paginate_query
from app.services.movements import register_movement

router = APIRouter(prefix="/tools", tags=["tools"], dependencies=[Depends(require_login)])


def _filtered_tools_query(db: Session, search: str, search_col: str):
    """Build a filtered query for tools. Shared by listing and export."""
    query = db.query(Tool).join(ToolType).filter(Tool.is_critical != -1)

    if search:
        like = f"%{search}%"
        if search_col == "name":
            query = query.filter(Tool.name.ilike(like))
        elif search_col == "origin_id":
            query = query.filter(Tool.origin_id.ilike(like))
        elif search_col == "tool_type":
            query = query.filter(ToolType.name.ilike(like))
        elif search_col == "location":
            query = query.filter(Tool.location.ilike(like))
        elif search_col == "status":
            s_lower = search.strip().lower()
            if any(kw in s_lower for kw in ["crít", "crit"]):
                query = query.filter(Tool.current_stock <= 0)
            elif any(kw in s_lower for kw in ["baixo", "low"]):
                query = query.filter(Tool.current_stock > 0, Tool.current_stock < Tool.min_stock)
            elif "ok" in s_lower:
                query = query.filter(Tool.current_stock >= Tool.min_stock)
            else:
                query = query.filter(False)
        else:  # "all"
            query = query.filter(
                (Tool.name.ilike(like)) | (ToolType.name.ilike(like))
            )

    return query


_TOOL_SORT_COLUMNS = {
    "id": Tool.id,
    "origin_id": Tool.origin_id,
    "name": Tool.name,
    "tool_type": ToolType.name,
    "location": Tool.location,
    "min_stock": Tool.min_stock,
    "max_stock": Tool.max_stock,
    "current_stock": Tool.current_stock,
    "unit_cost": Tool.unit_cost,
    "status": case(
        (Tool.current_stock <= 0, 0),
        (and_(Tool.current_stock > 0, Tool.current_stock < Tool.min_stock), 1),
        else_=2,
    ),
}


def _apply_tools_sort(query, sort_by: str, sort_dir: str):
    normalized_sort_by = sort_by if sort_by in _TOOL_SORT_COLUMNS else "name"
    normalized_sort_dir = "desc" if sort_dir == "desc" else "asc"

    sort_column = _TOOL_SORT_COLUMNS[normalized_sort_by]
    ordered = sort_column.desc() if normalized_sort_dir == "desc" else sort_column.asc()
    tie_breaker = Tool.id.asc() if normalized_sort_by != "id" else Tool.name.asc()
    return query.order_by(ordered, tie_breaker), normalized_sort_by, normalized_sort_dir


@router.get("/")
def tools_list(
    request: Request,
    search: str = Query("", alias="search"),
    search_col: str = Query("all", alias="search_col"),
    sort_by: str = Query("name", alias="sort_by"),
    sort_dir: str = Query("asc", alias="sort_dir"),
    page: int = Query(1, ge=1),
    per_page: int = Query(25, alias="per_page"),
    db: Session = Depends(get_db),
):
    """Display tools table with search and status highlighting."""
    tools_query, normalized_sort_by, normalized_sort_dir = _apply_tools_sort(
        _filtered_tools_query(db, search, search_col),
        sort_by,
        sort_dir,
    )
    tools, pagination = paginate_query(
        tools_query,
        base_path=request.url.path,
        params={
            "search": search,
            "search_col": search_col,
            "sort_by": normalized_sort_by,
            "sort_dir": normalized_sort_dir,
        },
        page=page,
        per_page=per_page,
        id_prefix="tools",
    )

    employees = db.query(Employee).order_by(Employee.name).all()
    machines = db.query(Machine).order_by(Machine.name).all()

    return request.app.state.templates.TemplateResponse(
        request=request,
        name="tools/index.html",
        context={
            "tools": tools,
            "employees": employees,
            "machines": machines,
            "search": search,
            "search_col": search_col,
            "current_sort_by": normalized_sort_by,
            "current_sort_dir": normalized_sort_dir,
            "page": pagination["page"],
            "per_page": pagination["per_page"],
            "pagination": pagination,
        },
    )


# ---- Column definitions for XLSX export (key -> label, extractor) ----
_STATUS_MAP = {"CRITICAL": "CRÍTICO", "LOW STOCK": "ESTOQUE BAIXO", "OK": "OK"}
_EXPORT_COLUMNS = {
    "id":            ("ID",                  lambda t: t.id),
    "origin_id":     ("ID TOTVS",            lambda t: t.origin_id),
    "name":          ("Nome da Ferramenta",   lambda t: t.name),
    "tool_type":     ("Tipo",                 lambda t: t.tool_type.name if t.tool_type else ""),
    "location":      ("Localização",          lambda t: t.location),
    "min_stock":     ("Mín.",                 lambda t: t.min_stock),
    "max_stock":     ("Máx.",                 lambda t: t.max_stock),
    "current_stock": ("Estoque",              lambda t: t.current_stock),
    "status":        ("Status",               lambda t: _STATUS_MAP.get(t.status, t.status)),
}
_ALL_COLUMN_KEYS = list(_EXPORT_COLUMNS.keys())


@router.get("/export/xlsx")
def export_tools_xlsx(
    request: Request,
    search: str = Query("", alias="search"),
    search_col: str = Query("all", alias="search_col"),
    sort_by: str = Query("name", alias="sort_by"),
    sort_dir: str = Query("asc", alias="sort_dir"),
    columns: list[str] = Query(None, alias="columns"),
    db: Session = Depends(get_db),
):
    """Export filtered tools browse to an XLSX file with selectable columns."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    # Sanitize columns against allowlist; fallback to all if empty/invalid
    if columns:
        selected = [c for c in columns if c in _EXPORT_COLUMNS]
    else:
        selected = _ALL_COLUMN_KEYS
    if not selected:
        selected = _ALL_COLUMN_KEYS

    tools_query, _, _ = _apply_tools_sort(_filtered_tools_query(db, search, search_col), sort_by, sort_dir)
    tools = tools_query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Estoque de Ferramentas"

    # Header row
    bold = Font(bold=True)
    for col_idx, key in enumerate(selected, 1):
        cell = ws.cell(row=1, column=col_idx, value=_EXPORT_COLUMNS[key][0])
        cell.font = bold

    # Data rows
    for row_idx, t in enumerate(tools, 2):
        for col_idx, key in enumerate(selected, 1):
            ws.cell(row=row_idx, column=col_idx, value=_EXPORT_COLUMNS[key][1](t))

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 3

    # Autofilter
    ws.auto_filter.ref = ws.dimensions

    # Write to buffer
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"estoque_ferramentas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/print/labels")
def print_labels(
    request: Request,
    search: str = Query("", alias="search"),
    search_col: str = Query("all", alias="search_col"),
    sort_by: str = Query("name", alias="sort_by"),
    sort_dir: str = Query("asc", alias="sort_dir"),
    db: Session = Depends(get_db),
):
    """Render a printable A4 page with labels (4×13 grid) for filtered tools."""
    tools_query, _, _ = _apply_tools_sort(_filtered_tools_query(db, search, search_col), sort_by, sort_dir)
    tools = tools_query.all()
    labels = [{"origin_id": t.origin_id or "", "name": t.name, "location": t.location or ""} for t in tools]

    # Chunk into pages of 52 labels (4 cols × 13 rows)
    per_page = 52
    pages = [labels[i:i + per_page] for i in range(0, len(labels), per_page)] if labels else []

    return request.app.state.templates.TemplateResponse(
        request=request,
        name="tools/labels_print.html",
        context={"pages": pages, "labels": labels},
    )


@router.get("/print/movement-form")
def print_movement_form(request: Request):
    """Render a printable offline movement form for use during system downtime."""
    return request.app.state.templates.TemplateResponse(
        request=request,
        name="tools/movement_form_print.html",
        context={
            "generated_on": datetime.now().strftime("%d/%m/%Y"),
            "rows": list(range(1, 12)),
        },
    )


@router.get("/print/loan-form")
def print_loan_form(request: Request):
    """Render a printable loan form for manual tool check-out and return."""
    return request.app.state.templates.TemplateResponse(
        request=request,
        name="tools/loan_form_print.html",
        context={
            "generated_on": datetime.now().strftime("%d/%m/%Y"),
            "rows": list(range(1, 12)),
        },
    )


@router.get("/create")
def tool_create_form(request: Request, db: Session = Depends(get_db)):
    tool_types = db.query(ToolType).order_by(ToolType.name).all()
    return request.app.state.templates.TemplateResponse(
        request=request,
        name="tools/create.html",
        context={"tool_types": tool_types},
    )


@router.post("/create")
async def tool_create(
    request: Request,
    db: Session = Depends(get_db),
):
    form = await request.form()
    name = form.get("name", "")
    origin_id = form.get("origin_id", "")
    tool_type_id = int(form.get("tool_type_id", 0))
    description = form.get("description", "")
    min_stock = int(form.get("min_stock", 0))
    max_stock = int(form.get("max_stock", 0))

    gaveta = form.get("gaveta", "")
    divisoria = form.get("divisoria", "")
    location = f"G{gaveta}D{divisoria}" if gaveta and divisoria else ""

    tool = Tool(
        name=name,
        origin_id=origin_id,
        tool_type_id=tool_type_id,
        description=description,
        location=location,
        min_stock=min_stock,
        max_stock=max_stock,
        current_stock=0,
    )
    db.add(tool)
    db.flush()  # get tool.id

    # Dynamic parameters
    param_names = form.getlist("param_name")
    param_values = form.getlist("param_value")
    for pn, pv in zip(param_names, param_values):
        if pn.strip():
            db.add(ToolParameter(tool_id=tool.id, parameter_name=pn.strip(), parameter_value=pv.strip()))

    db.commit()
    return RedirectResponse(url="/tools", status_code=303)


@router.post("/movement")
async def tool_movement(
    request: Request,
    db: Session = Depends(get_db),
):
    """Handle Add Stock (IN) or Remove Stock (OUT) via form submission."""
    form = await request.form()
    tool_id = int(form.get("tool_id", 0))
    movement_type = form.get("movement_type", "IN").upper()
    quantity = int(form.get("quantity", 0))
    notes = form.get("notes", "")
    category = form.get("category", "EMPRESTIMO").upper()
    raw_cost = form.get("unit_cost", "")
    unit_cost = float(raw_cost) if raw_cost else None

    employee_id = None
    machine_id = None
    if category == "REPOSICAO" and movement_type == "OUT":
        machine_id = int(form.get("machine_id", 0)) or None
    else:
        employee_id = int(form.get("employee_id", 0)) or None

    try:
        register_movement(
            db, tool_id, employee_id, movement_type, quantity, notes,
            category=category, machine_id=machine_id, unit_cost=unit_cost,
        )
    except ValueError:
        pass  # Silently redirect – in production add flash messages

    redirect_to = form.get("redirect_to", "/tools")
    # Only allow relative paths to prevent open redirect
    if not redirect_to.startswith("/"):
        redirect_to = "/tools"
    return RedirectResponse(url=redirect_to, status_code=303)


@router.get("/{tool_id}/modal")
def tool_detail_modal(
    tool_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    tool = db.query(Tool).filter(Tool.id == tool_id).first()
    if not tool:
        raise HTTPException(status_code=404, detail="Ferramenta não encontrada")

    tool_types = db.query(ToolType).order_by(ToolType.name).all()

    return request.app.state.templates.TemplateResponse(
        request=request,
        name="tools/detail_modal_content.html",
        context={
            "tool": tool,
            "tool_types": tool_types,
        },
    )


@router.get("/{tool_id}")
def tool_detail(
    tool_id: int,
    request: Request,
    ref: str = Query("", alias="ref"),
    ref_col: str = Query("", alias="ref_col"),
    db: Session = Depends(get_db),
):
    tool = db.query(Tool).filter(Tool.id == tool_id).first()
    if not tool:
        return RedirectResponse(url="/tools", status_code=303)

    tool_types = db.query(ToolType).order_by(ToolType.name).all()
    employees = db.query(Employee).order_by(Employee.name).all()
    machines = db.query(Machine).order_by(Machine.name).all()

    return request.app.state.templates.TemplateResponse(
        request=request,
        name="tools/detail.html",
        context={
            "tool": tool,
            "tool_types": tool_types,
            "employees": employees,
            "machines": machines,
            "ref": ref,
            "ref_col": ref_col,
        },
    )


@router.post("/{tool_id}/edit")
async def tool_edit(tool_id: int, request: Request, db: Session = Depends(get_db)):
    """Update tool details (does NOT change current_stock)."""
    tool = db.query(Tool).filter(Tool.id == tool_id).first()
    if not tool:
        return RedirectResponse(url="/tools", status_code=303)

    form = await request.form()
    tool.name = form.get("name", tool.name)
    tool.origin_id = form.get("origin_id", tool.origin_id)
    tool.tool_type_id = int(form.get("tool_type_id", tool.tool_type_id))
    tool.description = form.get("description", tool.description)
    tool.min_stock = int(form.get("min_stock", tool.min_stock))
    tool.max_stock = int(form.get("max_stock", tool.max_stock))
    tool.unit_cost = float(form.get("unit_cost", tool.unit_cost) or 0)
    tool.is_critical = 1 if form.get("is_critical") else 0
    tool.avg_lifespan_hours = float(form.get("avg_lifespan_hours", tool.avg_lifespan_hours) or 0)

    gaveta = form.get("gaveta", "")
    divisoria = form.get("divisoria", "")
    if gaveta and divisoria:
        tool.location = f"G{gaveta}D{divisoria}"
    elif not gaveta and not divisoria:
        tool.location = ""

    # Update dynamic parameters: remove old, add new
    db.query(ToolParameter).filter(ToolParameter.tool_id == tool.id).delete()
    param_names = form.getlist("param_name")
    param_values = form.getlist("param_value")
    for pn, pv in zip(param_names, param_values):
        if pn.strip():
            db.add(ToolParameter(tool_id=tool.id, parameter_name=pn.strip(), parameter_value=pv.strip()))

    db.commit()
    return RedirectResponse(url="/tools", status_code=303)


@router.post("/{tool_id}/delete")
async def tool_delete(tool_id: int, request: Request, db: Session = Depends(get_db)):
    """Deactivate a tool by marking current_stock = -1 and hiding it, or fully delete if no movements."""
    tool = db.query(Tool).filter(Tool.id == tool_id).first()
    if not tool:
        return RedirectResponse(url="/tools", status_code=303)

    from app.models import Movement
    has_movements = db.query(Movement).filter(Movement.tool_id == tool_id).count() > 0

    if has_movements:
        # Has history – soft delete: set inactive flag
        tool.is_critical = -1  # Convention: -1 = inactive/deleted
        tool.description = f"[DESATIVADA] {tool.description}"
        db.commit()
    else:
        # No history – safe to hard delete
        db.query(ToolParameter).filter(ToolParameter.tool_id == tool_id).delete()
        db.delete(tool)
        db.commit()

    return RedirectResponse(url="/tools", status_code=303)
