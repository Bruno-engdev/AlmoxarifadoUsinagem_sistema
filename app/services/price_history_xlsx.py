"""
Adapter de leitura do arquivo de histórico de preços do TOTVS.

Responsabilidade única: abrir o arquivo, detectar cabeçalho, normalizar tipos
(strings com zero à esquerda, datas dd/mm/yyyy, decimais com vírgula) e
retornar uma lista de :class:`ImportedPriceRow`. NÃO acessa banco de dados.

Formatos suportados:
- .xlsx (via openpyxl)
- .xml no formato "XML Spreadsheet 2003" (Workbook do Excel salvo como XML)

Cabeçalho esperado (case-insensitive, espaços ignorados):
    Produto, Numero, Dt.Entrega, Dt.Emissao, Fornecedor, Nome Fornec, Tipo,
    Item, Descricao, Unidade, Segunda UM, Quantidade, Preco KG, Prc Unitario,
    Ultimo Preco, Aliq. IPI, Observacoes, Numero da SC, Qtd.Entregue
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterator
from xml.etree import ElementTree as ET


# ---------------------------------------------------------------------------
# Public dataclasses
# ---------------------------------------------------------------------------

@dataclass
class ImportedPriceRow:
    """Linha normalizada do arquivo TOTVS, pronta para o serviço de import."""
    row_number: int
    origin_id: str
    numero_documento: str = ""
    data_entrega: date | None = None
    data_emissao: date | None = None
    fornecedor_codigo: str = ""
    fornecedor_nome: str = ""
    tipo: str = ""
    item: str = ""
    descricao_totvs: str = ""
    unidade: str = ""
    segunda_unidade: str = ""
    quantidade: Decimal | None = None
    preco_kg: Decimal | None = None
    preco_unitario: Decimal | None = None
    ultimo_preco: Decimal | None = None
    aliquota_ipi: Decimal | None = None
    observacoes: str = ""
    numero_sc: str = ""
    qtd_entregue: Decimal | None = None

    parse_errors: list[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        """Linha mínima utilizável: tem código de produto e preço unitário."""
        return (
            not self.parse_errors
            and bool(self.origin_id)
            and self.preco_unitario is not None
            and self.preco_unitario > 0
        )


# ---------------------------------------------------------------------------
# Header normalization + mapping
# ---------------------------------------------------------------------------

# Cada chave é o nome canônico interno; valor é a lista de variantes de
# cabeçalho aceitas no arquivo (após normalização: sem acento, lowercase,
# sem caracteres não-alfanuméricos).
HEADER_ALIASES: dict[str, list[str]] = {
    "origin_id":         ["produto", "codigo", "codproduto", "codigoproduto"],
    "numero_documento":  ["numero", "numerodocumento", "documento", "doc"],
    "data_entrega":      ["dtentrega", "dataentrega", "dtentr"],
    "data_emissao":      ["dtemissao", "dataemissao", "dtemis"],
    "fornecedor_codigo": ["fornecedor", "codfornec", "codigofornecedor"],
    "fornecedor_nome":   ["nomefornec", "nomefornecedor", "razaosocial"],
    "tipo":              ["tipo"],
    "item":              ["item"],
    "descricao_totvs":   ["descricao", "desc", "descproduto"],
    "unidade":           ["unidade", "um", "un"],
    "segunda_unidade":   ["segundaum", "segundaunidade", "um2"],
    "quantidade":        ["quantidade", "qtd", "qtde"],
    "preco_kg":          ["precokg", "prcoperkg", "prckg"],
    "preco_unitario":    ["prcunitario", "precounitario", "preco", "prcunit"],
    "ultimo_preco":      ["ultimopreco", "ultpreco"],
    "aliquota_ipi":      ["aliqipi", "ipi", "aliquotaipi"],
    "observacoes":       ["observacoes", "obs", "observacao"],
    "numero_sc":         ["numerodasc", "numerosc", "sc"],
    "qtd_entregue":      ["qtdentregue", "quantidadeentregue", "qtdentr"],
}


def _normalize_header(value: Any) -> str:
    """Lowercase, remove acentos, remove caracteres não-alfanuméricos."""
    if value is None:
        return ""
    s = str(value).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Za-z0-9]", "", s)
    return s.lower()


def _build_header_index(header_row: list[Any]) -> dict[str, int]:
    """Mapeia nome canônico → índice da coluna no header_row."""
    mapping: dict[str, int] = {}
    normalized = [_normalize_header(h) for h in header_row]
    for canonical, aliases in HEADER_ALIASES.items():
        for idx, header in enumerate(normalized):
            if header in aliases:
                mapping[canonical] = idx
                break
    return mapping


# ---------------------------------------------------------------------------
# Value normalizers
# ---------------------------------------------------------------------------

def _to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        # Excel devolve códigos numéricos como float (ex.: 13019.0); preserva
        # como inteiro em string. Zero à esquerda já foi perdido aqui — só
        # consegue ser preservado se o Excel armazenou como texto.
        return str(int(value))
    return str(value).strip()


def _to_origin_id(value: Any, expected_width: int = 6) -> str:
    """Normaliza código do produto: tira espaços e re-aplica zero à esquerda
    quando o valor veio como número (perdendo zeros)."""
    s = _to_str(value)
    if not s:
        return ""
    # Se for puramente numérico e menor que a largura esperada, completa zeros.
    if s.isdigit() and len(s) < expected_width:
        s = s.zfill(expected_width)
    return s


def _to_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    s = str(value).strip()
    if not s or s == "-":
        return None
    # Remove separador de milhar (ponto) quando há vírgula decimal.
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


# ---------------------------------------------------------------------------
# Row parser (independente da origem)
# ---------------------------------------------------------------------------

def _parse_row(row_number: int, raw: list[Any],
               header_index: dict[str, int]) -> ImportedPriceRow:
    def get(canonical: str) -> Any:
        idx = header_index.get(canonical)
        if idx is None or idx >= len(raw):
            return None
        return raw[idx]

    parsed = ImportedPriceRow(
        row_number=row_number,
        origin_id=_to_origin_id(get("origin_id")),
        numero_documento=_to_str(get("numero_documento")),
        data_entrega=_to_date(get("data_entrega")),
        data_emissao=_to_date(get("data_emissao")),
        fornecedor_codigo=_to_str(get("fornecedor_codigo")),
        fornecedor_nome=_to_str(get("fornecedor_nome")),
        tipo=_to_str(get("tipo")),
        item=_to_str(get("item")),
        descricao_totvs=_to_str(get("descricao_totvs")),
        unidade=_to_str(get("unidade")),
        segunda_unidade=_to_str(get("segunda_unidade")),
        quantidade=_to_decimal(get("quantidade")),
        preco_kg=_to_decimal(get("preco_kg")),
        preco_unitario=_to_decimal(get("preco_unitario")),
        ultimo_preco=_to_decimal(get("ultimo_preco")),
        aliquota_ipi=_to_decimal(get("aliquota_ipi")),
        observacoes=_to_str(get("observacoes")),
        numero_sc=_to_str(get("numero_sc")),
        qtd_entregue=_to_decimal(get("qtd_entregue")),
    )

    if not parsed.origin_id:
        parsed.parse_errors.append("Produto vazio")
    if parsed.preco_unitario is None:
        parsed.parse_errors.append("Prc Unitário inválido ou vazio")
    elif parsed.preco_unitario <= 0:
        parsed.parse_errors.append("Prc Unitário deve ser > 0")

    return parsed


# ---------------------------------------------------------------------------
# XLSX reader
# ---------------------------------------------------------------------------

def _iter_xlsx_rows(file_path: Path, sheet_name: str | None) -> Iterator[list[Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        for row in ws.iter_rows(values_only=True):
            yield list(row)
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# XML Spreadsheet 2003 reader
# ---------------------------------------------------------------------------

_XML_NS = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}


def _iter_xml_rows(file_path: Path, sheet_name: str | None) -> Iterator[list[Any]]:
    tree = ET.parse(str(file_path))
    root = tree.getroot()
    worksheets = root.findall("ss:Worksheet", _XML_NS)
    if not worksheets:
        return
    target = None
    if sheet_name:
        for ws in worksheets:
            if ws.attrib.get(f"{{{_XML_NS['ss']}}}Name") == sheet_name:
                target = ws
                break
    if target is None:
        target = worksheets[0]

    table = target.find("ss:Table", _XML_NS)
    if table is None:
        return

    for row in table.findall("ss:Row", _XML_NS):
        cells: list[Any] = []
        current_index = 0
        for cell in row.findall("ss:Cell", _XML_NS):
            # Suporta atributo ss:Index para pular colunas.
            idx_attr = cell.attrib.get(f"{{{_XML_NS['ss']}}}Index")
            if idx_attr:
                target_idx = int(idx_attr) - 1
                while current_index < target_idx:
                    cells.append(None)
                    current_index += 1
            data = cell.find("ss:Data", _XML_NS)
            if data is None:
                cells.append(None)
            else:
                dtype = data.attrib.get(f"{{{_XML_NS['ss']}}}Type", "String")
                text = data.text or ""
                if dtype == "Number":
                    try:
                        cells.append(float(text))
                    except ValueError:
                        cells.append(text)
                elif dtype == "DateTime":
                    try:
                        cells.append(datetime.fromisoformat(text))
                    except ValueError:
                        cells.append(text)
                else:
                    cells.append(text)
            current_index += 1
        yield cells


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def read_totvs_price_file(file_path: str | Path,
                          sheet_name: str | None = None) -> list[ImportedPriceRow]:
    """
    Lê o arquivo do TOTVS e devolve uma lista de linhas normalizadas.

    Detecta automaticamente .xlsx vs .xml. A primeira linha não vazia que
    contiver pelo menos um cabeçalho reconhecido é tratada como header; as
    seguintes são interpretadas como dados.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {path}")

    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        row_iter = _iter_xlsx_rows(path, sheet_name)
    elif suffix == ".xml":
        row_iter = _iter_xml_rows(path, sheet_name)
    else:
        raise ValueError(
            f"Formato não suportado: {suffix}. Use .xlsx ou .xml (XML Spreadsheet 2003)."
        )

    header_index: dict[str, int] = {}
    parsed_rows: list[ImportedPriceRow] = []

    for row_number, raw in enumerate(row_iter, start=1):
        if not raw or all(c is None or (isinstance(c, str) and not c.strip()) for c in raw):
            continue

        if not header_index:
            candidate = _build_header_index(raw)
            # Considera header se reconheceu pelo menos origin_id e preco_unitario.
            if "origin_id" in candidate and "preco_unitario" in candidate:
                header_index = candidate
            continue

        parsed_rows.append(_parse_row(row_number, raw, header_index))

    if not header_index:
        raise ValueError(
            "Cabeçalho do TOTVS não reconhecido. Esperado pelo menos as colunas "
            "'Produto' e 'Prc Unitário'."
        )

    return parsed_rows
